from __future__ import annotations

import csv
import gzip
import hashlib
import json
import pickle
import re
from collections.abc import Iterator
from io import BufferedReader
from pathlib import Path

import wslPath
from openpyxl import Workbook, load_workbook

###########################################################
# Input/Output
###########################################################


# =========================================================
# SAM/BAM
# =========================================================
def read_sam(path_of_sam: str | Path) -> Iterator[list]:
    with open(path_of_sam) as f:
        for line in f:
            yield line.strip().split("\t")


# =========================================================
# Pickle
# =========================================================


def load_pickle(file_path: Path):
    with open(file_path, "rb") as f:
        return pickle.load(f)


def save_pickle(data: object, file_path: Path) -> None:
    with open(file_path, "wb") as f:
        pickle.dump(data, f)


# =========================================================
# JSONL
# =========================================================


def read_jsonl(file_path: str | Path) -> Iterator[dict]:
    with open(file_path) as f:
        for line in f:
            yield json.loads(line)


def write_jsonl(data: list[dict] | Iterator[dict], file_path: str | Path) -> None:
    with open(file_path, "w") as f:
        for d in data:
            json_str = json.dumps(d)
            f.write(json_str + "\n")


# =========================================================
# Table (Excel, CSV)
# =========================================================


def read_xlsx(file_path: str | Path) -> list[dict[str, str]]:
    """Load data from an Excel file."""
    wb = load_workbook(filename=file_path)
    ws = wb.active

    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(element is None for element in row):  # Skip rows with all None values
            continue
        row_data = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
        records.append(row_data)

    return records


def read_csv(file_path: str | Path) -> list[dict[str, str]]:
    """Load data from a CSV file."""
    with open(file_path) as csvfile:
        header = [field.strip() for field in next(csv.reader(csvfile))]

        records = []
        for row in csv.reader(csvfile):
            if not row:  # Skip empty rows
                continue
            if all(element is None for element in row):  # Skip rows with all None values
                continue
            row_trimmed = [field.strip() for field in row]
            row_data = dict(zip(header, row_trimmed))
            records.append(row_data)

        return records


def write_xlsx(data: list[dict[str, str]], file_path: str | Path) -> None:
    # Create a workbook and a worksheet
    wb = Workbook()
    ws = wb.active

    if not data:
        wb.save(file_path)
        return

    # Initialize headers with the keys of the first dictionary, maintaining order
    headers = list(data[0].keys())

    # Check for new keys in the subsequent dictionaries and append them to the headers list
    for item in data[1:]:
        for key in item.keys():
            if key not in headers:
                headers.append(key)

    # Write the headers to the first row
    ws.append(headers)

    # Write the data to Excel
    for item in data:
        row = [item.get(header, "") for header in headers]
        ws.append(row)

    # Save the file
    wb.save(file_path)


# =========================================================
# FASTA / FASTQ
# =========================================================


def is_gzip_file(path_file: str | Path) -> bool:
    """Check if a file is a GZip compressed file."""
    try:
        with Path(path_file).open("rb") as f:
            return f.read(2) == b"\x1f\x8b"
    except OSError:
        return False


def read_lines(path_fastx: str | Path) -> Iterator[str]:
    if is_gzip_file(path_fastx):
        with gzip.open(path_fastx, "rt") as f:
            yield from [line.strip() for line in f]
    else:
        with open(path_fastx) as f:
            yield from [line.strip() for line in f]


def detect_fastx_format(path_fastx: str | Path) -> str:
    fastx_iterator = read_lines(path_fastx)
    lines = [next(fastx_iterator).strip() for _ in range(4)]

    if lines[0].startswith(">"):
        return "FASTA"
    elif lines[0].startswith("@") and lines[2].startswith("+"):
        return "FASTQ"
    raise ValueError("File format is unknown. The input does not appear to be in FASTA or FASTQ format.")


def read_fastq(path_fastx: str | Path) -> Iterator[dict]:
    """Read a FASTQ file and yield each record as a dictionary."""
    iterator = read_lines(path_fastx)
    try:
        while True:
            identifier = next(iterator).strip()
            sequence = next(iterator).strip()
            separator = next(iterator).strip()
            quality = next(iterator).strip()
            yield {"identifier": identifier, "sequence": sequence, "separator": separator, "quality": quality}
    except StopIteration:
        return


def read_fasta(path_fasta: str | Path) -> Iterator[dict]:
    """Read a FASTA file and yield each record as a dictionary."""
    iterator = read_lines(path_fasta)
    try:
        while True:
            identifier = next(iterator).strip()
            sequence = next(iterator).strip()
            yield {"identifier": identifier, "sequence": sequence}
    except StopIteration:
        return


def write_fasta(data: list[dict] | Iterator[dict], file_path: str | Path, is_gzip: bool = True) -> None:
    open_func = gzip.open if is_gzip else open
    mode = "wt" if is_gzip else "w"

    with open_func(file_path, mode) as f:
        for record in data:
            f.write(f"{record['identifier']}\n{record['sequence']}\n")


def write_fastq(data: list[dict] | Iterator[dict], file_path: str | Path, is_gzip: bool = True) -> None:
    open_func = gzip.open if is_gzip else open
    mode = "wt" if is_gzip else "w"

    with open_func(file_path, mode) as f:
        for record in data:
            f.write(f"{record['identifier']}\n{record['sequence']}\n{record['separator']}\n{record['quality']}\n")


###########################################################
# Load batch file
###########################################################


def determine_file_type(file_path: str) -> str | None:
    """Determine if the file is an Excel or CSV file. Raise error for other types."""
    file_extension = Path(file_path).suffix
    if file_extension in [".xlsx", ".xls"]:
        return "excel"
    elif file_extension == ".csv":
        return "csv"
    else:
        raise ValueError("The provided file must be either an Excel or CSV file.")


def load_batchfile(batchfile_path: str) -> list[dict[str, str]]:
    """Load data from either an Excel or CSV file."""
    file_type = determine_file_type(batchfile_path)
    if file_type == "excel":
        return read_xlsx(batchfile_path)
    elif file_type == "csv":
        return read_csv(batchfile_path)


###########################################################
# File hander
###########################################################


def count_newlines(filepath: str | Path) -> int:
    def read_in_chunks(file: BufferedReader, chunk_size: int = 2**16) -> Iterator[bytes, None, None]:
        """Get a Iterator that reads a file in chunks and yields each chunk."""
        while True:
            chunk = file.read(chunk_size)
            if not chunk:
                break
            yield chunk

    # Open the file in binary mode and count the newline characters
    with open(filepath, "rb") as file:
        count = sum(chunk.count(b"\n") for chunk in read_in_chunks(file))

    return count


def convert_to_posix(path: str) -> str:
    if wslPath.is_windows_path(path):
        path = wslPath.to_posix(path)
    return path


def sanitize_name(file_name: Path | str) -> str:
    """
    Sanitize the file name by replacing invalid characters on Windows OS with '-'
    """
    file_name = str(file_name).strip()
    if not file_name:
        raise ValueError("Provided name is empty or consists only of whitespace")
    forbidden_chars = r'[<>:"/\\|?*\x00-\x1F .]'

    return re.sub(forbidden_chars, "-", file_name)


###########################################################
# Cache control hash
###########################################################


def cache_control_hash(tempdir: Path, path_allele: str | Path) -> None:
    """
    Generate a hash value from the content of the provided file and cache it.

    Parameters:
    - tempdir: Path object, the temporary directory path.
    - path_allele: Path object, the path to the fasta file.
    """
    # Read the content of the file.
    content = Path(path_allele).read_text()
    # Calculate the hash of the content.
    content_hash = hashlib.sha256(content.encode()).hexdigest()
    # Define the path to save the hash.
    path_cache_hash = Path(tempdir, "cache", "hash.txt")
    # Ensure the directory exists.
    path_cache_hash.parent.mkdir(parents=True, exist_ok=True)
    # Save the hash to the file.
    path_cache_hash.write_text(content_hash)
